import tkinter as tk
from tkinter import ttk, filedialog, colorchooser, scrolledtext, messagebox
from tkinterdnd2 import TkinterDnD, DND_FILES
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import os

class TransMergeUI(TkinterDnD.Tk):
    def __init__(self):
        super().__init__()
        self.title("多Sheet译文合并替换工具 | 精准Key匹配")
        self.geometry("950x700")
        self.configure(bg='#f5f5f5')

        # 全局变量
        self.src_path = tk.StringVar()
        self.new_path = tk.StringVar()
        self.target_col = tk.StringVar(value="English")
        self.highlight_color = "#ADD8E6"
        self.highlight_rgb = "ADD8E6"

        self.init_ui()
        self.setup_drag_drop()

    def init_ui(self):
        # 标题
        title_label = tk.Label(self, text="Excel翻译合并工具", 
                              font=("Arial", 16, "bold"), bg='#f5f5f5', fg='#333')
        title_label.pack(pady=10)

        # 1. 原文件（被修改）
        frame1 = ttk.LabelFrame(self, text="📄 原文件（被替换译文）", padding=10)
        frame1.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Entry(frame1, textvariable=self.src_path, width=85).grid(row=0, column=0, padx=5, pady=5)
        ttk.Button(frame1, text="选择文件", command=self.select_src, width=12).grid(row=0, column=1, padx=5)
        
        # 拖拽区域
        self.src_drop_label = tk.Label(frame1, text="📁 拖拽文件到这里", 
                                       bg='#e8f4fd', relief=tk.SUNKEN, width=20, height=2,
                                       font=("Arial", 9))
        self.src_drop_label.grid(row=0, column=2, padx=5)

        # 2. 新译文文件
        frame2 = ttk.LabelFrame(self, text="📄 新译文文件（来源）", padding=10)
        frame2.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Entry(frame2, textvariable=self.new_path, width=85).grid(row=0, column=0, padx=5, pady=5)
        ttk.Button(frame2, text="选择文件", command=self.select_new, width=12).grid(row=0, column=1, padx=5)
        
        # 拖拽区域
        self.new_drop_label = tk.Label(frame2, text="📁 拖拽文件到这里", 
                                       bg='#e8f4fd', relief=tk.SUNKEN, width=20, height=2,
                                       font=("Arial", 9))
        self.new_drop_label.grid(row=0, column=2, padx=5)

        # 3. 语种列 + 颜色选择
        frame3 = ttk.LabelFrame(self, text="⚙️ 配置选项", padding=10)
        frame3.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(frame3, text="🎯 要替换的语种列名：").grid(row=0, column=0, sticky=tk.W, padx=5)
        
        # 【修改】下拉列表代替输入框
        self.language_options = [
            "English", "Japanese", "Korean", "Russian", "German", 
            "Chinese(Hong Kong)", "Chinese(Taiwan)", "Thai", "Vietnamese", 
            "Indonesian", "Portuguese", "French", "Spanish", "Turkish", 
            "Arabic", "Malay", "UZ", "UR"
        ]
        self.target_col_combo = ttk.Combobox(frame3, textvariable=self.target_col, 
                                            values=self.language_options, width=28, state="readonly")
        self.target_col_combo.grid(row=0, column=1, padx=5, pady=5)
        self.target_col_combo.current(0)  # 默认选中"English"（索引0）
        
        ttk.Button(frame3, text="🎨 选择高亮颜色", command=self.choose_color, width=15).grid(row=0, column=2, padx=10, pady=5)
        
        # 颜色预览
        self.color_label = tk.Label(frame3, text="  颜色预览  ", fg="black", bg=self.highlight_color, 
                                   width=12, relief=tk.RAISED, font=("Arial", 9, "bold"))
        self.color_label.grid(row=0, column=3, padx=5)

        # 4. 开始按钮
        frame4 = ttk.Frame(self, padding=10)
        frame4.pack(fill=tk.X)
        start_btn = ttk.Button(frame4, text="🚀 开始合并替换", command=self.start_merge, 
                              width=20, style="Accent.TButton")
        start_btn.pack(pady=10)
        
        # 自定义按钮样式
        style = ttk.Style()
        style.configure("Accent.TButton", font=("Arial", 11, "bold"))

        # 5. 清空日志按钮
        ttk.Button(frame4, text="🧹 清空日志", command=self.clear_log, width=15).pack(pady=5)

        # 6. 日志输出
        frame5 = ttk.LabelFrame(self, text="📋 运行日志", padding=10)
        frame5.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        self.log_text = scrolledtext.ScrolledText(frame5, wrap=tk.WORD, 
                                                  font=("Consolas", 9), height=20)
        self.log_text.pack(fill=tk.BOTH, expand=True)

        # 底部状态栏
        self.status_bar = tk.Label(self, text="就绪", bd=1, relief=tk.SUNKEN, anchor=tk.W, bg='#e0e0e0')
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

    # ========== 拖拽功能 ==========
    def setup_drag_drop(self):
        """设置拖拽功能"""
        self.src_drop_label.drop_target_register(DND_FILES)
        self.src_drop_label.dnd_bind('<<Drop>>', self.drop_src_file)
        
        self.new_drop_label.drop_target_register(DND_FILES)
        self.new_drop_label.dnd_bind('<<Drop>>', self.drop_new_file)
        
        # 悬停效果
        self.src_drop_label.bind('<Enter>', lambda e: self.src_drop_label.config(bg='#d0e8f7'))
        self.src_drop_label.bind('<Leave>', lambda e: self.src_drop_label.config(bg='#e8f4fd'))
        self.new_drop_label.bind('<Enter>', lambda e: self.new_drop_label.config(bg='#d0e8f7'))
        self.new_drop_label.bind('<Leave>', lambda e: self.new_drop_label.config(bg='#e8f4fd'))

    def drop_src_file(self, event):
        """处理原文件拖拽"""
        file_path = self._clean_drop_path(event.data)
        if file_path and file_path.endswith('.xlsx'):
            self.src_path.set(file_path)
            self.log(f"✅ 拖拽选择原文件：{os.path.basename(file_path)}")
            self.status_bar.config(text=f"原文件已选择: {os.path.basename(file_path)}")
        else:
            self.log("❌ 拖拽的文件不是有效的Excel文件（.xlsx）")

    def drop_new_file(self, event):
        """处理新文件拖拽"""
        file_path = self._clean_drop_path(event.data)
        if file_path and file_path.endswith('.xlsx'):
            self.new_path.set(file_path)
            self.log(f"✅ 拖拽选择新译文文件：{os.path.basename(file_path)}")
            self.status_bar.config(text=f"新文件已选择: {os.path.basename(file_path)}")
        else:
            self.log("❌ 拖拽的文件不是有效的Excel文件（.xlsx）")

    def _clean_drop_path(self, path):
        """清理拖拽路径"""
        path = path.strip()
        if path.startswith('{') and path.endswith('}'):
            path = path[1:-1]
        return path if os.path.exists(path) else None

    # ========== 选文件 ==========
    def select_src(self):
        path = filedialog.askopenfilename(filetypes=[("Excel文件", "*.xlsx")])
        if path:
            self.src_path.set(path)
            self.log(f"✅ 已选择原文件：{os.path.basename(path)}")
            self.status_bar.config(text=f"原文件已选择: {os.path.basename(path)}")

    def select_new(self):
        path = filedialog.askopenfilename(filetypes=[("Excel文件", "*.xlsx")])
        if path:
            self.new_path.set(path)
            self.log(f"✅ 已选择新译文文件：{os.path.basename(path)}")
            self.status_bar.config(text=f"新文件已选择: {os.path.basename(path)}")

    # ========== 选颜色 ==========
    def choose_color(self):
        color = colorchooser.askcolor(title="选择替换内容高亮颜色", initialcolor=self.highlight_color)
        if color[1]:
            self.highlight_color = color[1]
            self.highlight_rgb = color[1].lstrip("#")
            self.color_label.config(bg=self.highlight_color)
            self.log(f"✅ 已选择高亮颜色：{self.highlight_color}")
            self.status_bar.config(text=f"高亮颜色已设置: {self.highlight_color}")

    # ========== 清空日志 ==========
    def clear_log(self):
        self.log_text.delete(1.0, tk.END)
        self.log("✅ 日志已清空")
        self.status_bar.config(text="日志已清空", bg='#e0e0e0')

    # ========== 日志 ==========
    def log(self, msg):
        time_str = datetime.now().strftime("%H:%M:%S")
        line = f"[{time_str}] {msg}\n"
        self.log_text.insert(tk.END, line)
        self.log_text.see(tk.END)
        self.update()

    # ========== 核心工具方法 ==========
    @staticmethod
    def get_exact_cell_value(cell):
        """完全原始取值：不做任何去除、修剪、转义，严格全等"""
        if cell is None or cell.value is None:
            return ""
        return str(cell.value)

    def get_column_index_by_header(self, ws, header_name):
        """根据表头名找列下标"""
        for col_idx, cell in enumerate(ws[1]):
            val = self.get_exact_cell_value(cell)
            if val == header_name:
                return col_idx
        return -1

    # ========== 主合并逻辑 ==========
    def start_merge(self):
        src_file = self.src_path.get().strip()
        new_file = self.new_path.get().strip()
        col_name = self.target_col.get().strip()

        # 基础校验
        if not src_file or not os.path.exists(src_file):
            self.log("❌ 原文件不存在或未选择！")
            self.status_bar.config(text="错误: 原文件不存在", fg="red")
            return
        if not new_file or not os.path.exists(new_file):
            self.log("❌ 新译文文件不存在或未选择！")
            self.status_bar.config(text="错误: 新译文文件不存在", fg="red")
            return
        if not col_name:
            self.log("❌ 请输入需要替换的语种列名！")
            self.status_bar.config(text="错误: 未指定语种列名", fg="red")
            return

        self.log("="*60)
        self.log("🚀 开始执行译文合并替换...")
        self.log(f"📄 原文件：{src_file}")
        self.log(f"📄 新译文：{new_file}")
        self.log(f"🎯 目标替换列：{col_name}")
        self.log(f"🎨 高亮颜色：{self.highlight_color}")
        self.log("="*60)

        try:
            self.status_bar.config(text="正在处理...请稍候", fg="blue")
            self.update()

            # 记录开始时间
            start_time = datetime.now()

            # 加载工作簿
            self.log("\n📂 正在加载Excel文件...")
            wb_src = load_workbook(src_file)
            wb_new = load_workbook(new_file)
            self.log("✅ 文件加载成功")

            # 高亮样式
            fill_style = PatternFill(
                fill_type="solid",
                start_color=self.highlight_rgb,
                end_color=self.highlight_rgb
            )

            # 统计信息
            total_sheets = 0
            total_updates = 0
            total_mismatches = 0
            total_same = 0  # 【新增】统计相同译文数量
            error_log = []

            # 遍历所有页签
            for sheet_name in wb_src.sheetnames:
                total_sheets += 1
                self.log(f"\n{'='*50}")
                self.log(f"📄 处理工作表：【{sheet_name}】")
                self.log(f"{'='*50}")
                
                if sheet_name not in wb_new.sheetnames:
                    self.log(f"⚠️  新文件无此工作表，跳过：{sheet_name}")
                    continue

                ws_src = wb_src[sheet_name]
                ws_new = wb_new[sheet_name]

                # 查找目标语种列
                col_src_idx = self.get_column_index_by_header(ws_src, col_name)
                col_new_idx = self.get_column_index_by_header(ws_new, col_name)
                if col_src_idx == -1 or col_new_idx == -1:
                    self.log(f"❌ 当前表格未找到列名：'{col_name}'，跳过")
                    continue

                self.log(f"✅ 找到目标列：第 {col_src_idx+1} 列（{col_name}）")

                # 构建新文件Key字典
                new_key_dict = {}
                new_rows_count = 0
                for row in ws_new.iter_rows(min_row=2):
                    key_val = self.get_exact_cell_value(row[0])
                    if key_val:
                        new_key_dict[key_val] = row
                        new_rows_count += 1

                self.log(f"📊 新文件中找到 {new_rows_count} 个有效Key")

                # 遍历原文件所有行
                sheet_updates = 0
                sheet_mismatches = 0
                sheet_same = 0  # 【新增】当前工作表相同译文数量
                src_rows_count = 0
                
                for row in ws_src.iter_rows(min_row=2):
                    src_rows_count += 1
                    key_src = self.get_exact_cell_value(row[0])

                    # 空key跳过
                    if not key_src:
                        continue

                    # key不匹配
                    if key_src not in new_key_dict:
                        sheet_mismatches += 1
                        total_mismatches += 1
                        error_msg = f"❌ Key不匹配 | 工作表:{sheet_name} | Key：「{key_src}」| 新文件无此Key"
                        self.log(error_msg)
                        error_log.append(error_msg)
                        continue

                    # 取出新译文
                    new_row = new_key_dict[key_src]
                    new_text = self.get_exact_cell_value(new_row[col_new_idx])

                    # 新译文为空，不替换
                    if new_text.strip() == "":
                        self.log(f"ℹ️  Key「{key_src[:30]}...」新译文为空，跳过替换")
                        continue

                    # 【新增】获取原译文并比较
                    target_cell = row[col_src_idx]
                    old_text = self.get_exact_cell_value(target_cell)
                    
                    # 译文相同，跳过（不替换不上色）
                    if old_text == new_text:
                        sheet_same += 1
                        total_same += 1
                        continue

                    # 译文不同，替换+高亮
                    target_cell.value = new_text
                    target_cell.fill = fill_style
                    sheet_updates += 1
                    total_updates += 1
                    
                    # 显示简短日志
                    if len(key_src) > 30:
                        display_key = key_src[:27] + "..."
                    else:
                        display_key = key_src
                    
                    self.log(f"✅ [{sheet_updates}] Key「{display_key}」替换成功")

                self.log(f"\n📊 工作表「{sheet_name}」处理完成：")
                self.log(f"   - 原文件行数: {src_rows_count}")
                self.log(f"   - ✅ 成功更新: {sheet_updates} 条")
                self.log(f"   - ⚠️  Key不匹配: {sheet_mismatches} 条")
                self.log(f"   - ℹ️  译文相同: {sheet_same} 条（已跳过）")

            # 保存输出文件
            out_dir = os.path.dirname(src_file)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            out_name = f"译文合并结果_{timestamp}.xlsx"
            out_path = os.path.join(out_dir, out_name)
            
            self.log(f"\n💾 正在保存输出文件...")
            wb_src.save(out_path)
            self.log(f"✅ 输出文件已保存：{out_path}")

            # 保存详细错误日志
            if error_log:
                error_log_path = os.path.join(out_dir, f"Key不匹配日志_{timestamp}.txt")
                with open(error_log_path, "w", encoding="utf-8") as f:
                    f.write("="*60 + "\n")
                    f.write("Key不匹配详细日志\n")
                    f.write(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write(f"原文件: {src_file}\n")
                    f.write(f"新文件: {new_file}\n")
                    f.write("="*60 + "\n\n")
                    for i, error in enumerate(error_log, 1):
                        f.write(f"{i}. {error}\n")
                self.log(f"📄 Key不匹配详细日志已保存：{error_log_path}")

            # 保存运行日志
            log_path = os.path.join(out_dir, f"合并运行日志_{timestamp}.txt")
            with open(log_path, "w", encoding="utf-8") as f:
                f.write(self.log_text.get("1.0", tk.END))
            self.log(f"📄 运行日志已保存：{log_path}")

            # 计算耗时
            end_time = datetime.now()
            duration = (end_time - start_time).total_seconds()

            # 最终统计
            self.log(f"\n{'='*60}")
            self.log("🎉 全部处理完成！")
            self.log(f"{'='*60}")
            self.log(f"📊 处理统计：")
            self.log(f"   - 处理工作表: {total_sheets} 个")
            self.log(f"   - ✅ 成功更新: {total_updates} 条")
            self.log(f"   - ⚠️  Key不匹配: {total_mismatches} 条")
            self.log(f"   - ℹ️  译文相同: {total_same} 条（已跳过）")
            self.log(f"   - ⏱️  耗时: {duration:.2f} 秒")
            self.log(f"{'='*60}")
            self.log(f"📁 输出文件：{out_path}")
            
            if error_log:
                self.log(f"⚠️  注意：有 {total_mismatches} 个Key不匹配，请查看错误日志文件")
            
            self.status_bar.config(text=f"✅ 完成! 更新{total_updates}条, 相同{total_same}条, 不匹配{total_mismatches}条", fg="green")
            
            messagebox.showinfo("✅ 完成", 
                              f"处理完成！\n\n"
                              f"✅ 成功更新: {total_updates} 条\n"
                              f"⚠️  Key不匹配: {total_mismatches} 条\n"
                              f"ℹ️  译文相同: {total_same} 条（已跳过）\n\n"
                              f"📁 输出文件: {out_path}\n\n"
                              f"⏱️  耗时: {duration:.2f} 秒")

        except Exception as e:
            error_msg = f"❌ 程序异常：{str(e)}"
            self.log(error_msg)
            import traceback
            traceback.print_exc()
            self.status_bar.config(text="❌ 错误: 处理失败", fg="red")
            messagebox.showerror("❌ 错误", f"处理过程中发生错误:\n{str(e)}")

if __name__ == "__main__":
    try:
        app = TransMergeUI()
        app.mainloop()
    except ImportError as e:
        print("缺少依赖包，请安装:")
        print("pip install tkinterdnd2 openpyxl")
        input("按回车键退出...")